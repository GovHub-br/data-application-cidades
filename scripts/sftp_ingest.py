import io
import logging
import os
import re
import stat
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path

import paramiko
import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy import types as sa_types
from dotenv import load_dotenv

load_dotenv(Path(__file__).parent.parent / ".env")

SOCKET_ERRORS = (
    paramiko.SSHException,
    EOFError,
    OSError,
    ConnectionResetError,
)
_CONN_ERROR_STRINGS = (
    "garbage packet", "eof", "connection reset",
    "broken pipe", "timed out", "channel closed", "socket is closed",
)

_LOG_PATH = (
    Path(__file__).parent
    / f"sftp_errors_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(_LOG_PATH, encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

SFTP_HOST     = os.getenv("SFTP_HOST")
SFTP_PORT     = int(os.getenv("SFTP_PORT") or 22)
SFTP_USER     = os.getenv("SFTP_USER")
SFTP_PASSWORD = os.getenv("SFTP_PASSWORD")

DB_HOST     = os.getenv("DB_DW_HOST_MCID")
DB_PORT     = os.getenv("DB_DW_PORT_MCID")
DB_USER     = os.getenv("DB_DW_USER_MCID")
DB_PASSWORD = os.getenv("DB_DW_PASSWORD_MCID")
DB_NAME     = os.getenv("DB_DW_DBNAME_MCID")
DB_SCHEMA   = "sftp"
INGEST_LOG  = "_ingest_log"

SUPPORTED_EXTENSIONS = {".csv", ".txt", ".xlsx"}
IGNORED_EXTENSIONS   = {
    ".bashrc", ".bash_logout", ".bash_history", ".profile",
    ".mkshrc", ".viminfo", ".filepart", ".rpt",
}
CHUNK_SIZE            = 200_000
ENCODING_SCAN_BYTES   = 5 * 1024 * 1024  # 5 MB — suficiente para detectar encoding
ENCODINGS             = ["utf-8", "latin-1", "cp1252"]
DELIMITERS       = [";", ",", "|", "\t"]
TABLE_NAME_LIMIT = 57  # postgres cria _tablename como array type: 1 + 57 + 5(_0001) = 63

def normalize_name(name: str) -> str:
    name = name.lower()
    name = re.sub(r"[^\w]", "_", name)
    name = re.sub(r"_+", "_", name)
    return name.strip("_")


def deduplicate_columns(columns: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    result = []
    for col in columns:
        if col in seen:
            seen[col] += 1
            result.append(f"{col}_{seen[col]}")
        else:
            seen[col] = 0
            result.append(col)
    return result


def build_table_name(sftp_path: str, inner_filename: str | None = None) -> str:
    parts = Path(sftp_path).parts
    segments = [normalize_name(p) for p in parts if p not in (".", "..")]

    if inner_filename:
        # Ignora o nome do zip; usa pastas + stem do arquivo interno
        folder_segs = segments[:-1]
        inner_stem = normalize_name(Path(inner_filename).stem)
        name = "_".join(folder_segs + [inner_stem])
        if len(name) <= TABLE_NAME_LIMIT:
            return name
        # Muito longo: mantém prefixo das pastas + fim alinhado do stem interno
        # (preserva a parte mais descritiva, que fica no final do nome)
        prefix = "_".join(folder_segs)
        available = TABLE_NAME_LIMIT - len(prefix) - 1  # -1 pelo _ separador
        if available <= 0:
            return prefix[:TABLE_NAME_LIMIT]
        raw_suffix = inner_stem[-available:]
        us = raw_suffix.find("_")
        suffix = (
            raw_suffix[us + 1:] if us >= 0 and raw_suffix[us + 1:]
            else raw_suffix.lstrip("_")
        )
        return f"{prefix}_{suffix}"
    else:
        segments[-1] = normalize_name(Path(parts[-1]).stem)
        name = "_".join(segments)
        return name[:TABLE_NAME_LIMIT]


def resolve_table_name(base: str, registry: dict[str, int]) -> str:
    if base not in registry:
        registry[base] = 0
        return base
    registry[base] += 1
    return f"{base}_{registry[base]:04d}"


def format_size(size_bytes: int) -> str:
    value: float = size_bytes
    for unit in ["B", "KB", "MB", "GB", "TB"]:
        if value < 1024:
            return f"{value:.1f} {unit}"
        value /= 1024
    return f"{value:.1f} PB"

def collect_sftp_files(sftp: paramiko.SFTPClient) -> list[dict]:
    entries = []

    def walk(path: str):
        try:
            items = sftp.listdir_attr(path)
        except Exception as e:
            log.error("erro ao listar %s: %s", path, e)
            return
        for item in items:
            full_path = f"{path}/{item.filename}"
            ext = Path(item.filename).suffix.lower()
            if item.st_mode is not None and stat.S_ISDIR(item.st_mode):
                walk(full_path)
            elif ext in IGNORED_EXTENSIONS or item.filename.startswith("."):
                continue
            else:
                entries.append({
                    "path": full_path,
                    "size": item.st_size or 0,
                    "ext": ext,
                    "is_zip": ext == ".zip",
                })

    walk(".")
    return entries


def sort_files(entries: list[dict]) -> list[dict]:
    non_zip = sorted([e for e in entries if not e["is_zip"]], key=lambda x: x["size"])
    zips    = sorted([e for e in entries if e["is_zip"]],     key=lambda x: x["size"])
    return non_zip + zips


def download_to_tempfile(sftp: paramiko.SFTPClient, remote_path: str) -> str:
    """Baixa para disco e retorna o path — não carrega o arquivo inteiro na RAM."""
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=Path(remote_path).suffix)
    try:
        sftp.getfo(remote_path, tmp)
        tmp.flush()
    finally:
        tmp.close()
    return tmp.name

def detect_csv_format(file_path: str) -> tuple[str, str]:
    """Detecta encoding nos primeiros 5 MB e delimitador nos primeiros 8192 bytes."""
    enc = "latin-1"
    for candidate in ENCODINGS[:-1]:  # utf-8, cp1252 — latin-1 é fallback garantido
        try:
            with open(file_path, "rb") as f:
                raw = f.read(ENCODING_SCAN_BYTES)
            raw.decode(candidate)
            enc = candidate
            break
        except UnicodeDecodeError:
            continue

    with open(file_path, "rb") as f:
        sample = f.read(8192)
    try:
        first_line = sample.decode(enc).splitlines()[0]
    except Exception:
        first_line = sample.decode("latin-1").splitlines()[0]

    sep = ";"
    for candidate_sep in DELIMITERS:
        if first_line.count(candidate_sep) > 0:
            sep = candidate_sep
            break

    return enc, sep


def _copy_df_to_pg(df: pd.DataFrame, table_name: str, schema: str, engine) -> None:
    """Bulk insert via PostgreSQL COPY — 10-20x mais rápido que INSERT."""
    buf = io.StringIO()
    df.to_csv(buf, index=False, header=False, na_rep="\\N")
    buf.seek(0)
    raw_conn = engine.raw_connection()
    try:
        with raw_conn.cursor() as cur:
            cols = ", ".join(f'"{c}"' for c in df.columns)
            cur.copy_expert(
                f"COPY \"{schema}\".\"{table_name}\" ({cols}) "
                f"FROM STDIN WITH (FORMAT CSV, NULL '\\N')",
                buf,
            )
        raw_conn.commit()
    except Exception:
        raw_conn.rollback()
        raise
    finally:
        raw_conn.close()


def stream_csv_from_file(file_path: str, table_name: str, engine) -> int:
    """Lê CSV/TXT em chunks de um arquivo em disco."""
    enc, sep = detect_csv_format(file_path)
    log.info("    encoding=%s, sep=%s", enc, repr(sep))

    total = 0
    first = True
    for chunk in pd.read_csv(file_path, sep=sep, encoding=enc, chunksize=CHUNK_SIZE,
                              low_memory=False, on_bad_lines="skip"):
        chunk = chunk.loc[:, ~chunk.columns.str.match(r"^Unnamed: \d+$")]
        chunk.columns = deduplicate_columns(
            [normalize_name(str(c)) for c in chunk.columns]
        )
        if first:
            text_dtype = {c: sa_types.Text() for c in chunk.columns}
            chunk.head(0).to_sql(table_name, engine, schema=DB_SCHEMA,
                                 if_exists="replace", index=False, dtype=text_dtype)
            first = False
        _copy_df_to_pg(chunk, table_name, DB_SCHEMA, engine)
        total += len(chunk)
    return total


def stream_xlsx_from_file(file_path: str, table_name: str, engine) -> int:
    """Lê XLSX linha a linha com openpyxl read_only — não carrega a planilha inteira."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    total = 0
    multi_sheet = len(wb.sheetnames) > 1

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = iter(ws.rows)
        header_cells = list(next(rows_iter))

        # Filtra apenas colunas com header real (não-None, não vazio).
        # openpyxl read_only retorna células formatadas além dos dados — ignoramos todas.
        valid_indices = [
            i for i, c in enumerate(header_cells)
            if c.value is not None and str(c.value).strip() != ""
        ]
        if not valid_indices:
            log.warning("    aba '%s' sem headers, pulando", sheet_name)
            continue

        headers = deduplicate_columns(
            [normalize_name(str(header_cells[i].value)) for i in valid_indices]
        )

        buffer = []
        first = True
        for row in rows_iter:
            row_list = list(row)
            buffer.append([
                row_list[i].value if i < len(row_list) else None
                for i in valid_indices
            ])
            if len(buffer) >= CHUNK_SIZE:
                df = pd.DataFrame(buffer, columns=headers)
                if multi_sheet:
                    df["_sheet"] = sheet_name
                if first:
                    text_dtype = {c: sa_types.Text() for c in df.columns}
                    df.head(0).to_sql(table_name, engine, schema=DB_SCHEMA,
                                      if_exists="replace", index=False,
                                      dtype=text_dtype)
                    first = False
                _copy_df_to_pg(df, table_name, DB_SCHEMA, engine)
                total += len(df)
                buffer = []

        if buffer:
            df = pd.DataFrame(buffer, columns=headers)
            if multi_sheet:
                df["_sheet"] = sheet_name
            if first:
                text_dtype = {c: sa_types.Text() for c in df.columns}
                df.head(0).to_sql(table_name, engine, schema=DB_SCHEMA,
                                  if_exists="replace", index=False,
                                  dtype=text_dtype)
            _copy_df_to_pg(df, table_name, DB_SCHEMA, engine)
            total += len(df)

    wb.close()
    return total


def write_bytes_to_tempfile(data: bytes, suffix: str) -> str:
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(data)
        return tmp.name


def load_file_from_disk(file_path: str, ext: str, table_name: str, engine) -> int:
    if ext in (".csv", ".txt"):
        return stream_csv_from_file(file_path, table_name, engine)
    if ext == ".xlsx":
        return stream_xlsx_from_file(file_path, table_name, engine)
    log.warning("extensão não suportada: %s", ext)
    return -1

def ensure_ingest_log(engine) -> None:
    with engine.begin() as conn:
        conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS {DB_SCHEMA}.{INGEST_LOG} (
                sftp_key       TEXT PRIMARY KEY,
                table_name     TEXT,
                rows_inserted  INTEGER,
                processed_at   TIMESTAMP DEFAULT NOW()
            )
        """))


def is_processed(engine, sftp_key: str) -> bool:
    with engine.connect() as conn:
        result = conn.execute(
            text(f"SELECT 1 FROM {DB_SCHEMA}.{INGEST_LOG} WHERE sftp_key = :k"),
            {"k": sftp_key},
        )
        return result.fetchone() is not None


def is_zip_processed(engine, zip_path: str) -> bool:
    """Checa se o zip já tem ao menos um arquivo interno no log — evita baixar de novo."""
    with engine.connect() as conn:
        result = conn.execute(
            text(
                f"SELECT 1 FROM {DB_SCHEMA}.{INGEST_LOG}"
                f" WHERE sftp_key LIKE :prefix LIMIT 1"
            ),
            {"prefix": zip_path + "::%"},
        )
        return result.fetchone() is not None


def mark_processed(engine, sftp_key: str, table_name: str, rows: int) -> None:
    with engine.begin() as conn:
        conn.execute(
            text(f"""
                INSERT INTO {DB_SCHEMA}.{INGEST_LOG}
                    (sftp_key, table_name, rows_inserted)
                VALUES (:k, :t, :r)
                ON CONFLICT (sftp_key) DO UPDATE
                    SET rows_inserted = EXCLUDED.rows_inserted,
                        processed_at  = NOW()
            """),
            {"k": sftp_key, "t": table_name, "r": rows},
        )

def get_engine():
    url = (f"postgresql+psycopg2://{DB_USER}:{DB_PASSWORD}"
           f"@{DB_HOST}:{DB_PORT}/{DB_NAME}")
    return create_engine(url)


def ensure_schema(engine) -> None:
    with engine.begin() as conn:
        conn.execute(text(f"CREATE SCHEMA IF NOT EXISTS {DB_SCHEMA}"))

def _process_zip(tmp_path: str, sftp_path: str,
                 engine, registry: dict[str, int]) -> None:
    try:
        zf = zipfile.ZipFile(tmp_path)
    except zipfile.BadZipFile as e:
        log.error("zip inválido %s: %s", sftp_path, e)
        return

    inner_files = [
        f for f in zf.namelist()
        if Path(f).suffix.lower() in SUPPORTED_EXTENSIONS
        and not Path(f).name.startswith(".")
    ]
    if not inner_files:
        log.warning("nenhum arquivo suportado dentro de %s", sftp_path)
        return

    for inner in inner_files:
        sftp_key   = f"{sftp_path}::{inner}"
        inner_ext  = Path(inner).suffix.lower()
        base_name  = build_table_name(sftp_path, inner)
        table_name = resolve_table_name(base_name, registry)

        if is_processed(engine, sftp_key):
            log.info("  ↳ %s já processado, pulando", inner)
            continue

        log.info("  ↳ %s → %s.%s", inner, DB_SCHEMA, table_name)
        inner_tmp = write_bytes_to_tempfile(zf.read(inner), inner_ext)
        try:
            rows = load_file_from_disk(inner_tmp, inner_ext, table_name, engine)
            if rows >= 0:
                mark_processed(engine, sftp_key, table_name, rows)
                log.info("    ✓ %d linhas inseridas", rows)
        except Exception as e:
            log.error("erro em %s::%s: %s", sftp_path, inner, e)
        finally:
            os.unlink(inner_tmp)


def _process_flat(entry: dict, sftp: paramiko.SFTPClient,
                  engine, registry: dict[str, int]) -> None:
    path = entry["path"]
    ext  = entry["ext"]

    if ext not in SUPPORTED_EXTENSIONS:
        log.warning("extensão não suportada, pulando: %s", path)
        return

    if is_processed(engine, path):
        log.info("  já processado, pulando: %s", path)
        return

    base_name  = build_table_name(path)
    table_name = resolve_table_name(base_name, registry)
    log.info("  → %s.%s", DB_SCHEMA, table_name)

    tmp_path = download_to_tempfile(sftp, path)
    try:
        rows = load_file_from_disk(tmp_path, ext, table_name, engine)
        if rows >= 0:
            mark_processed(engine, path, table_name, rows)
            log.info("  ✓ %d linhas inseridas", rows)
    except Exception as e:
        log.error("erro em %s: %s", path, e)
    finally:
        os.unlink(tmp_path)


def process_entry(entry: dict, sftp: paramiko.SFTPClient,
                  engine, registry: dict[str, int]) -> None:
    log.info("\n→ %s (%s)", entry["path"], format_size(entry["size"]))

    if entry["is_zip"]:
        if is_zip_processed(engine, entry["path"]):
            log.info("  zip já processado, pulando")
            return
        tmp_path = download_to_tempfile(sftp, entry["path"])
        try:
            _process_zip(tmp_path, entry["path"], engine, registry)
        finally:
            os.unlink(tmp_path)
    else:
        _process_flat(entry, sftp, engine, registry)

def _is_conn_error(e: Exception) -> bool:
    return isinstance(e, SOCKET_ERRORS) or any(
        s in str(e).lower() for s in _CONN_ERROR_STRINGS
    )


def connect_sftp() -> tuple[paramiko.Transport, paramiko.SFTPClient]:
    if not SFTP_HOST:
        raise ValueError("SFTP_HOST não encontrada no .env")
    transport = paramiko.Transport((SFTP_HOST, SFTP_PORT))
    transport.connect(username=SFTP_USER, password=SFTP_PASSWORD)
    transport.set_keepalive(30)  # keepalive a cada 30s para evitar timeout do servidor
    sftp = paramiko.SFTPClient.from_transport(transport)
    if sftp is None:
        raise RuntimeError("Falha ao abrir sessão SFTP")
    return transport, sftp


def main():
    if not SFTP_PASSWORD:
        raise ValueError("SFTP_PASSWORD não encontrada no .env")

    log.info("Log: %s", _LOG_PATH)
    log.info("Conectando ao SFTP %s@%s:%s ...", SFTP_USER, SFTP_HOST, SFTP_PORT)
    transport, sftp = connect_sftp()

    log.info("Coletando lista de arquivos...")
    entries = collect_sftp_files(sftp)
    ordered = sort_files(entries)

    total_size    = sum(e["size"] for e in ordered)
    non_zip_count = sum(1 for e in ordered if not e["is_zip"])
    zip_count     = sum(1 for e in ordered if e["is_zip"])
    log.info("%d arquivos | %s", len(ordered), format_size(total_size))
    log.info("  %d não-zip (primeiro) | %d zip (depois)", non_zip_count, zip_count)

    engine = get_engine()
    ensure_schema(engine)
    ensure_ingest_log(engine)
    registry: dict[str, int] = {}

    for entry in ordered:
        try:
            process_entry(entry, sftp, engine, registry)
        except Exception as e:
            if not _is_conn_error(e):
                log.error("erro inesperado em %s: %s", entry["path"], e)
                continue
            log.warning("Conexão perdida (%s — %s), reconectando...", type(e).__name__, e)
            try:
                sftp.close()
                transport.close()
            except Exception:
                pass
            try:
                transport, sftp = connect_sftp()
            except Exception as conn_e:
                log.error("Falha ao reconectar: %s. Abortando.", conn_e)
                break
            log.info("Reconectado. Retentando %s ...", entry["path"])
            try:
                process_entry(entry, sftp, engine, registry)
            except Exception as e2:
                log.error("erro após reconexão em %s: %s", entry["path"], e2)

    sftp.close()
    transport.close()

    error_count = sum(
        1 for line in _LOG_PATH.read_text(encoding="utf-8").splitlines()
        if " ERROR " in line
    )
    log.info("Concluído. Erros: %d → %s", error_count, _LOG_PATH)


if __name__ == "__main__":
    main()
