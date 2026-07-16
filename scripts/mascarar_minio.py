# scripts/mascarar_minio.py

"""
Mascaramento de PII (dados de pessoa física) na camada raw/ do data lake (MinIO).

Percorre os objetos de raw/, detecta colunas sensíveis pelo header e mascara os valores,
sobrescrevendo o objeto no lugar (o raw deixa de conter PII).

Técnica (decisão do projeto):
  - Identificadores (CPF, NIS) -> token HMAC-SHA256 determinístico (mesmo valor gera o mesmo
    token em todas as bases; irreversível; preserva join/contagem de distintos).
  - Quasi-identificadores (nome de PF, endereço, CEP, data de nascimento) -> redação (valor fixo).
  - CEP/endereço só são mascarados quando o arquivo tem algum indicador de PF (CPF/NIS/
    nascimento/nome-PF); em bases PJ/empreendimento o CEP é preservado.

Segurança do dado não-alvo: o arquivo é lido e reescrito em transporte latin-1 (byte a byte),
então qualquer coluna não mascarada é gravada byte-idêntica, independentemente do encoding real
do arquivo. Apenas as colunas-alvo são substituídas por texto ASCII (token/redação). O encoding
real é detectado só para interpretar corretamente os NOMES das colunas (matching).

Idempotência: objetos já mascarados recebem a tag `masked=true`; execuções seguintes os pulam
(evita duplo-HMAC). Use --force para reprocessar.

Auditoria/linhagem: cada execução grava um parquet em audit/masking/execution_id=<uuid>/ no
MinIO e registra cada arquivo em sftp._masking_log no Postgres.

Por padrão roda em DRY-RUN (não sobrescreve raw/; grava a prévia em masked_dryrun/). Use --apply
para efetivar a sobrescrita.
"""

import argparse
import csv
import hashlib
import hmac
import io
import json
import logging
import os
import re
import sys
import tempfile
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
import psycopg2
from boto3.s3.transfer import TransferConfig
from dotenv import load_dotenv
from psycopg2.extras import Json

from lake_utils import (
    MDB_ENCODING,
    MDB_EXT,
    baixar,
    criar_cliente_minio,
    detectar_dialeto,
    detectar_encoding,
    md5_arquivo,
    mdb_contar,
    mdb_disponivel,
    mdb_header,
    mdb_tabelas,
    norm_header,
    sample_bytes,
)

load_dotenv()

MINIO_ENDPOINT   = os.environ["MINIO_ENDPOINT"]
MINIO_ACCESS_KEY = os.environ["MINIO_ACCESS_KEY"]
MINIO_SECRET_KEY = os.environ["MINIO_SECRET_KEY"]
MINIO_BUCKET     = os.environ["MINIO_BUCKET"]

PG_HOST     = os.environ["DB_DW_HOST_MCID"]
PG_PORT     = int(os.environ.get("DB_DW_PORT_MCID", 5432))
PG_USER     = os.environ["DB_DW_USER_MCID"]
PG_PASSWORD = os.environ["DB_DW_PASSWORD_MCID"]
PG_DBNAME   = os.environ["DB_DW_DBNAME_MCID"]

SCHEMA        = os.environ.get("SFTP_SCHEMA", "sftp")
CONTROL_TABLE = "_masking_log"

RAW_PREFIX     = os.environ.get("MASKING_PREFIX", "raw/")
DRYRUN_PREFIX  = "masked_dryrun/"
AUDIT_PREFIX   = "audit/masking/"

HMAC_SECRET = os.environ.get("MASKING_HMAC_SECRET", "").encode("utf-8")
TOKEN_LEN   = int(os.environ.get("MASKING_TOKEN_LEN", 16))
REDACTION   = os.environ.get("MASKING_REDACTION", "***")

# /tmp costuma ser tmpfs pequeno; os Base_PF_FGTS têm ~2 GB e o processamento mantém
# original + mascarado em disco ao mesmo tempo (~4,5 GB). Aponte para um disco com espaço.
TMPDIR = os.environ.get("MASKING_TMPDIR") or None
if TMPDIR:
    os.makedirs(TMPDIR, exist_ok=True)

SUPPORTED_TABULAR = {".csv", ".txt"}
SUPPORTED_EXCEL   = {".xlsx"}
SUPPORTED_MDB     = MDB_EXT          # .mdb/.accdb — só LEITURA (ver _analisar_mdb)
UNSUPPORTED       = {".xls", ".zip"}

# csv pode ter campos grandes (linhas longas de bases bancárias)
csv.field_size_limit(2**31 - 1)

# Padrões de detecção de colunas sensíveis (do mapeamento do schema sftp)
P_CPF   = re.compile(r"cpf")
# NIS/PIS/PASEP/NIT são o mesmo número de identificação do trabalhador (identificador de PF)
P_NIS   = re.compile(r"(^|_)nis(_|$)|nu_nis|num_nis|(^|_)pis(_|$)|pasep|(^|_)nit(_|$)")
P_CEP   = re.compile(r"cep")
P_ENDER = re.compile(
    r"endereco|logradouro|(^|_)rua(_|$)|bairro|complemento|"
    r"num_?casa|numero_?casa|(^|_)quadra(_|$)|(^|_)lote(_|$)"
)
# não-endereços que casariam por acidente: "objetivo_complemento" (rótulo de programa),
# "ic_benef_sit_rua" (flag indicadora de situação de rua)
P_ENDER_EXC = re.compile(r"objetivo|sit_rua|(^|_)ic(_|$)")
P_NASC  = re.compile(r"nascimento|dt_?nasc|data_?nasc|dat_nasc")
P_NOME_INC = re.compile(
    r"mutuario|beneficiario|comprador|titular|proponente|responsavel|"
    r"conjuge|dependente|completo"
)
P_NOME_EXC = re.compile(
    r"empreendimento|municipio|(^|_)uf(_|$)|agente|banco|entidade|orgao|"
    r"logradouro|bairro|arquivo|razao|social|programa|modalidade|situacao|"
    r"fantasia|projeto|obra|construtora|incorporadora|"
    # colunas com papel (mutuario/beneficiario/titular...) mas que não são NOME:
    # identificadores PJ, códigos, valores, flags e datas
    r"cnpj|cpf|sexo|(^|_)tipo(_|$)|(^|_)vr(_|$)|valor|prest|parcela|"
    r"(^|_)qt(_|$)|(^|_)ic(_|$)|(^|_)dt(_|$)|(^|_)mulher(_|$)|pdc|pcd|objetivo"
)

# categorias que indicam presença de PF no arquivo (habilitam mascaramento de CEP/endereço)
_PF_INDICATOR_CATS = {"cpf", "nis", "nascimento", "nome"}

_LOG_FILE = (
    Path(__file__).parent
    / f"mascarar_minio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
)
_formatter = logging.Formatter("%(asctime)s %(levelname)s %(message)s", datefmt="%H:%M:%S")
logging.root.setLevel(logging.INFO)
for _h in (logging.FileHandler(_LOG_FILE, encoding="utf-8"), logging.StreamHandler(sys.stderr)):
    _h.setFormatter(_formatter)
    logging.root.addHandler(_h)
log = logging.getLogger(__name__)


# Infra: conexões
def _conn_str() -> str:
    return (
        f"host={PG_HOST} port={PG_PORT} dbname={PG_DBNAME} "
        f"user={PG_USER} password={PG_PASSWORD}"
    )


def _cliente_minio():
    return criar_cliente_minio(MINIO_ENDPOINT, MINIO_ACCESS_KEY, MINIO_SECRET_KEY)


_TRANSFER_CONFIG = TransferConfig(
    multipart_chunksize=8 * 1024 * 1024,
    max_concurrency=2,
)


def _criar_control_table(conn_str: str) -> None:
    with psycopg2.connect(conn_str) as conn:
        with conn.cursor() as cur:
            cur.execute(f"CREATE SCHEMA IF NOT EXISTS {SCHEMA};")
            cur.execute(f"""
                CREATE TABLE IF NOT EXISTS {SCHEMA}.{CONTROL_TABLE} (
                    id                  SERIAL PRIMARY KEY,
                    execution_id        TEXT,
                    minio_key           TEXT NOT NULL,
                    file_name           TEXT,
                    source_hash         TEXT,
                    masked_hash         TEXT,
                    masked_columns      JSONB,
                    registros_total     BIGINT,
                    registros_alterados BIGINT,
                    status              TEXT,
                    error_message       TEXT,
                    created_at          TIMESTAMPTZ DEFAULT NOW(),
                    UNIQUE (minio_key, source_hash)
                );
            """)
            cur.execute(f"""
                CREATE INDEX IF NOT EXISTS idx_masking_log_status
                ON {SCHEMA}.{CONTROL_TABLE} (status);
            """)
            conn.commit()
    log.info("Tabela de controle %s.%s garantida.", SCHEMA, CONTROL_TABLE)


def _carregar_masked_hashes(conn_str: str) -> set:
    with psycopg2.connect(conn_str) as conn:
        with conn.cursor() as cur:
            cur.execute(f"""
                SELECT masked_hash FROM {SCHEMA}.{CONTROL_TABLE}
                WHERE status = 'masked' AND masked_hash IS NOT NULL
            """)
            return {row[0] for row in cur.fetchall()}


def _registrar_control(conn_str: str, row: dict) -> None:
    with psycopg2.connect(conn_str) as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                INSERT INTO {SCHEMA}.{CONTROL_TABLE}
                    (execution_id, minio_key, file_name, source_hash, masked_hash,
                     masked_columns, registros_total, registros_alterados, status, error_message)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (minio_key, source_hash) DO UPDATE SET
                    execution_id        = EXCLUDED.execution_id,
                    masked_hash         = EXCLUDED.masked_hash,
                    masked_columns      = EXCLUDED.masked_columns,
                    registros_total     = EXCLUDED.registros_total,
                    registros_alterados = EXCLUDED.registros_alterados,
                    status              = EXCLUDED.status,
                    error_message       = EXCLUDED.error_message,
                    created_at          = NOW()
                """,
                (
                    row["execution_id"], row["file"], row["file_name"], row["hash_before"],
                    row["hash_after"], Json(row["masked_columns"]), row["registros_total"],
                    row["registros_alterados"], row["status"], row["error_message"],
                ),
            )
            conn.commit()


# Mascaramento de valores
def _hmac_token(valor: str) -> str:
    if valor is None or valor.strip() == "":
        return valor
    dig = hmac.new(HMAC_SECRET, valor.strip().encode("utf-8"), hashlib.sha256).hexdigest()
    return dig[:TOKEN_LEN]


def _redigir(valor: str) -> str:
    if valor is None or valor.strip() == "":
        return valor
    return REDACTION


# Detecção de header / colunas sensíveis
def _categoria(norm: str) -> Optional[str]:
    """Categoria base da coluna (sem aplicar a regra condicional de CEP/endereço)."""
    if P_CPF.search(norm):
        return "cpf"
    if P_NIS.search(norm):
        return "nis"
    if P_NASC.search(norm):
        return "nascimento"
    if P_NOME_INC.search(norm) and not P_NOME_EXC.search(norm):
        return "nome"
    if norm == "nome":
        return "nome_bare"
    if P_CEP.search(norm):
        return "cep"
    if P_ENDER.search(norm) and not P_ENDER_EXC.search(norm):
        return "endereco"
    return None


def classificar(header: List[str], real_encoding: str) -> Tuple[List[dict], bool]:
    """
    Retorna (targets, has_pf_indicator).
    targets: [{idx, column, category, action}] já com a regra condicional aplicada.
    header vem lido em latin-1; reinterpretamos cada célula no encoding real p/ o matching.
    """
    normed: List[Tuple[int, str, str]] = []  # (idx, original_header, norm)
    for idx, cell in enumerate(header):
        try:
            texto = cell.encode("latin-1", "surrogateescape").decode(real_encoding, "replace")
        except Exception:
            texto = cell
        normed.append((idx, cell, norm_header(texto)))

    cats = {idx: _categoria(n) for idx, _, n in normed}
    has_pf = any(c in _PF_INDICATOR_CATS for c in cats.values())

    targets: List[dict] = []
    for idx, original, _ in normed:
        cat = cats[idx]
        if cat is None:
            continue
        if cat in ("cpf", "nis"):
            action = "hmac"
        elif cat in ("nascimento", "nome"):
            action = "redact"
        elif cat == "nome_bare":
            if not has_pf:
                continue
            cat, action = "nome", "redact"
        elif cat in ("cep", "endereco"):
            if not has_pf:  # PJ/empreendimento -> preserva
                continue
            action = "redact"
        else:
            continue
        targets.append(
            {"idx": idx, "column": original, "category": cat, "action": action}
        )
    return targets, has_pf


# Processamento CSV/TXT (streaming, byte-preserving via latin-1)
def _mascarar_tabular(
    src_path: str, dst_path: str, delim: str, lineterm: str, fully_quoted: bool,
    real_encoding: str,
) -> Tuple[List[dict], bool, int, int]:
    """Retorna (targets, has_pf, registros_total, registros_alterados)."""
    quoting = csv.QUOTE_ALL if fully_quoted else csv.QUOTE_MINIMAL
    total = alterados = 0
    targets: List[dict] = []
    has_pf = False

    with open(src_path, "r", encoding="latin-1", newline="") as fin, \
         open(dst_path, "w", encoding="latin-1", newline="") as fout:
        reader = csv.reader(fin, delimiter=delim, quotechar='"')
        writer = csv.writer(
            fout, delimiter=delim, quotechar='"', quoting=quoting, lineterminator=lineterm
        )

        try:
            header = next(reader)
        except StopIteration:
            return targets, has_pf, 0, 0

        targets, has_pf = classificar(header, real_encoding)
        writer.writerow(header)
        if not targets:
            # sem colunas sensíveis: nada a fazer (o chamador trata como skip_no_pii)
            return targets, has_pf, 0, 0

        idx_action = [(t["idx"], t["action"]) for t in targets]
        for row in reader:
            total += 1
            row_alterada = False
            for idx, action in idx_action:
                if idx < len(row) and row[idx] is not None and row[idx].strip() != "":
                    row[idx] = _hmac_token(row[idx]) if action == "hmac" else _redigir(row[idx])
                    row_alterada = True
            if row_alterada:
                alterados += 1
            writer.writerow(row)

    return targets, has_pf, total, alterados


def _verificar_roundtrip_tabular(
    src_path: str, dst_path: str, delim: str, total_esperado: int
) -> None:
    """Garante que nº de linhas/colunas do header foi preservado."""
    def _header_e_linhas(path):
        with open(path, "r", encoding="latin-1", newline="") as f:
            reader = csv.reader(f, delimiter=delim, quotechar='"')
            header = next(reader, [])
            n = sum(1 for _ in reader)
        return len(header), n

    ncols_src, _ = _header_e_linhas(src_path)
    ncols_dst, n_dst = _header_e_linhas(dst_path)
    if ncols_src != ncols_dst:
        raise ValueError(f"round-trip: colunas do header divergem ({ncols_src} != {ncols_dst})")
    if n_dst != total_esperado:
        raise ValueError(f"round-trip: nº de linhas divergem ({n_dst} != {total_esperado})")


# Processamento XLSX
def _xlsx_tem_alvo(src_path: str) -> Tuple[bool, bool]:
    """Pré-scan barato dos headers em modo read_only (streaming, sem carregar o DOM).

    load_workbook completo materializa TODAS as células como objetos na RAM (~0,5-1 KB
    por célula); fazer isso só para descobrir que o arquivo não tem PII é desperdício —
    e a maioria dos xlsx do lake não tem. Retorna (tem_alvo, has_pf).
    """
    import openpyxl

    wb = openpyxl.load_workbook(src_path, read_only=True)
    try:
        tem_alvo = False
        has_pf_any = False
        for ws in wb.worksheets:
            first = next(ws.iter_rows(values_only=True), None)
            if first is None:
                continue
            header = [str(c) if c is not None else "" for c in first]
            targets, has_pf = classificar(header, "utf-8")
            has_pf_any = has_pf_any or has_pf
            if targets:
                tem_alvo = True
        return tem_alvo, has_pf_any
    finally:
        wb.close()


def _mascarar_xlsx(src_path: str, dst_path: str) -> Tuple[List[dict], bool, int, int, bool]:
    """Retorna (targets, has_pf, registros_total, registros_alterados, has_formulas).

    Só deve ser chamada quando _xlsx_tem_alvo() indicou PII — carrega o workbook
    completo (DOM em RAM) para poder reescrever células preservando o restante.

    has_formulas: True se alguma célula não-alvo contém fórmula. Fórmulas são preservadas
    no arquivo de saída (nunca são tratadas como alvo de mascaramento), mas o openpyxl não
    recalcula: o valor em cache dessas células se perde ao salvar e só é restaurado quando o
    arquivo é reaberto (e resalvo) num Excel real. Leitores programáticos (pandas, data_only)
    verão None nessas células até lá — sinalizado na auditoria para quem for consumir o dado.
    """
    import openpyxl

    wb = openpyxl.load_workbook(src_path)
    total = alterados = 0
    all_targets: List[dict] = []
    has_pf_any = False
    has_formulas = False

    for ws in wb.worksheets:
        rows = ws.iter_rows()
        try:
            header_cells = next(rows)
        except StopIteration:
            continue
        header = [str(c.value) if c.value is not None else "" for c in header_cells]
        targets, has_pf = classificar(header, "utf-8")
        has_pf_any = has_pf_any or has_pf
        idx_action = [(t["idx"], t["action"]) for t in targets]
        idx_alvo = {t["idx"] for t in targets}
        if targets:
            all_targets.extend({**t, "sheet": ws.title} for t in targets)
        for row in rows:
            if targets:
                total += 1
            row_alterada = False
            for idx, cell in enumerate(row):
                if cell.data_type == "f" and idx not in idx_alvo:
                    has_formulas = True
            for idx, action in idx_action:
                if idx < len(row):
                    cell = row[idx]
                    val = cell.value
                    if val is not None and str(val).strip() != "":
                        cell.value = _hmac_token(str(val)) if action == "hmac" else _redigir(str(val))
                        row_alterada = True
            if row_alterada:
                alterados += 1

    wb.save(dst_path)
    return all_targets, has_pf_any, total, alterados, has_formulas


# Análise de .mdb (Access) — LEITURA APENAS
def _analisar_mdb(src_path: str) -> Tuple[List[dict], bool, int]:
    """Varre as tabelas do .mdb procurando colunas sensíveis. Retorna (targets, has_pf, linhas).

    NÃO mascara: mdbtools é read-only e não existe forma de reescrever um .mdb sem Java
    (Jackcess/UCanAccess). Esta função só responde "tem PII?" — se tiver, o chamador falha
    explicitamente, porque gravar PII silenciosamente no lake seria pior que um erro visível.

    Hoje as 4 famílias de .mdb do lake (AO_1/2/3, CCI_CCA, AF) são dados de empreendimento/obra
    e analítico agregado: sem CPF/nome/nascimento. O que existe é Gênero, Entidades.CGC (PJ) e
    tab_empreendimentos.txt_logradouro (endereço do EMPREENDIMENTO) — este último é preservado
    pela regra condicional de CEP/endereço, que só dispara com indicador de PF na tabela.
    """
    targets: List[dict] = []
    has_pf_any = False
    total = 0
    for tabela in mdb_tabelas(src_path):
        header = mdb_header(src_path, tabela)
        if not header:
            continue
        n = mdb_contar(src_path, tabela)
        if n > 0:
            total += n
        # mdb-export já entrega UTF-8, então o encoding aqui é conhecido
        t, has_pf = classificar(header, MDB_ENCODING)
        has_pf_any = has_pf_any or has_pf
        targets.extend({**x, "table": tabela} for x in t)
    return targets, has_pf_any, total


# S3 helpers
def _listar_objetos(s3, prefix: str):
    pag = s3.get_paginator("list_objects_v2")
    for page in pag.paginate(Bucket=MINIO_BUCKET, Prefix=prefix):
        for obj in page.get("Contents", []):
            yield obj["Key"], obj["Size"]


def _ja_mascarado(s3, key: str) -> bool:
    try:
        tags = s3.get_object_tagging(Bucket=MINIO_BUCKET, Key=key).get("TagSet", [])
        return any(t["Key"] == "masked" and t["Value"] == "true" for t in tags)
    except Exception:
        return False


def _subir(s3, path: str, key: str) -> None:
    with open(path, "rb") as f:
        s3.upload_fileobj(f, MINIO_BUCKET, key, Config=_TRANSFER_CONFIG)


def _tag_mascarado(s3, key: str, execution_id: str, masked_hash: str) -> None:
    s3.put_object_tagging(
        Bucket=MINIO_BUCKET,
        Key=key,
        Tagging={"TagSet": [
            {"Key": "masked", "Value": "true"},
            {"Key": "execution_id", "Value": execution_id},
            {"Key": "masked_hash", "Value": masked_hash},
        ]},
    )


# Processamento de um objeto
def _novo_registro(execution_id: str, key: str) -> dict:
    return {
        "execution_id": execution_id,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "bucket": MINIO_BUCKET,
        "file": key,
        "file_name": key.rsplit("/", 1)[-1],
        "file_format": key.rsplit(".", 1)[-1].lower() if "." in key.rsplit("/", 1)[-1] else "",
        "encoding": None,
        "delimiter": None,
        "has_pf_indicator": False,
        "has_formulas": False,
        "masked_columns": [],
        "registros_total": 0,
        "registros_alterados": 0,
        "hash_before": None,
        "hash_after": None,
        "status": None,
        "error_message": None,
        "duration_s": 0.0,
    }


def processar_objeto(
    s3, conn_str: str, key: str, execution_id: str, apply: bool, masked_hashes: set,
) -> dict:
    t0 = time.time()
    rec = _novo_registro(execution_id, key)
    ext = os.path.splitext(key)[1].lower()

    if ext in UNSUPPORTED:
        rec["status"] = "skipped_unsupported"
        rec["duration_s"] = round(time.time() - t0, 2)
        return rec

    src = dst = None
    try:
        sample = sample_bytes(s3, MINIO_BUCKET, key)
        real_encoding = detectar_encoding(sample)
        rec["encoding"] = real_encoding

        if ext in SUPPORTED_TABULAR:
            dialeto = detectar_dialeto(sample, real_encoding)
            if dialeto is None:
                rec["status"] = "skipped_no_header"
                rec["duration_s"] = round(time.time() - t0, 2)
                return rec
            delim, lineterm, fully_quoted = dialeto
            rec["delimiter"] = delim

            src = baixar(s3, MINIO_BUCKET, key, ext, TMPDIR)
            rec["hash_before"] = md5_arquivo(src)
            if rec["hash_before"] in masked_hashes:
                rec["status"] = "skipped_already"
                return rec

            dst = tempfile.NamedTemporaryFile(delete=False, suffix=ext, dir=TMPDIR).name
            targets, has_pf, total, alterados = _mascarar_tabular(
                src, dst, delim, lineterm, fully_quoted, real_encoding
            )
            rec.update(
                has_pf_indicator=has_pf,
                masked_columns=[
                    {k: t[k] for k in ("column", "category", "action")} for t in targets
                ],
                registros_total=total,
                registros_alterados=alterados,
            )
            if not targets:
                rec["status"] = "skipped_no_pii"
                return rec

            _verificar_roundtrip_tabular(src, dst, delim, total)

        elif ext in SUPPORTED_MDB:
            # .mdb é somente-leitura (mdbtools não escreve): aqui só verificamos se há PII.
            if not mdb_disponivel():
                raise RuntimeError(
                    "mdbtools não encontrado no PATH — necessário para ler .mdb "
                    "(instale o pacote 'mdbtools')."
                )
            src = baixar(s3, MINIO_BUCKET, key, ext, TMPDIR)
            rec["hash_before"] = md5_arquivo(src)
            if rec["hash_before"] in masked_hashes:
                rec["status"] = "skipped_already"
                return rec

            targets, has_pf, total = _analisar_mdb(src)
            rec.update(
                has_pf_indicator=has_pf,
                masked_columns=[
                    {k: t[k] for k in ("column", "category", "action")} for t in targets
                ],
                registros_total=total,
            )
            if not targets:
                rec["status"] = "skipped_no_pii"
                return rec

            # Tem PII e não há como reescrever .mdb — falhar alto em vez de fingir que mascarou.
            cols = ", ".join(f"{t['table']}.{t['column']}" for t in targets[:5])
            rec["status"] = "error"
            rec["error_message"] = (
                f"PII encontrada em .mdb ({len(targets)} coluna(s): {cols}) — mdbtools é "
                "read-only e não há como reescrever o arquivo. Tratar à parte (converter e "
                "descartar o .mdb, ou usar Jackcess/UCanAccess via Java)."
            )[:500]
            log.error("✗ %s: %s", key, rec["error_message"])
            return rec

        elif ext in SUPPORTED_EXCEL:
            src = baixar(s3, MINIO_BUCKET, key, ext, TMPDIR)
            rec["hash_before"] = md5_arquivo(src)
            if rec["hash_before"] in masked_hashes:
                rec["status"] = "skipped_already"
                return rec
            tem_alvo, has_pf_scan = _xlsx_tem_alvo(src)
            if not tem_alvo:
                rec["has_pf_indicator"] = has_pf_scan
                rec["status"] = "skipped_no_pii"
                return rec
            dst = tempfile.NamedTemporaryFile(delete=False, suffix=ext, dir=TMPDIR).name
            targets, has_pf, total, alterados, has_formulas = _mascarar_xlsx(src, dst)
            rec.update(
                has_pf_indicator=has_pf,
                masked_columns=[
                    {k: t[k] for k in ("column", "category", "action")} for t in targets
                ],
                registros_total=total,
                registros_alterados=alterados,
                has_formulas=has_formulas,
            )
            if not targets:
                rec["status"] = "skipped_no_pii"
                return rec
            if has_formulas:
                log.warning(
                    "%s contém fórmulas em colunas não mascaradas — valores em cache serão "
                    "perdidos até reabrir/resalvar num Excel real (leitura programática pode "
                    "ver None nessas células).",
                    key,
                )
        else:
            rec["status"] = "skipped_unsupported"
            return rec

        rec["hash_after"] = md5_arquivo(dst)

        if apply:
            _subir(s3, dst, key)
            _tag_mascarado(s3, key, execution_id, rec["hash_after"])
            rec["status"] = "masked"
        else:
            preview_key = DRYRUN_PREFIX + key[len(RAW_PREFIX):] if key.startswith(RAW_PREFIX) else DRYRUN_PREFIX + key
            _subir(s3, dst, preview_key)
            rec["status"] = "dry_run"

        return rec

    except Exception as e:  # noqa: BLE001
        rec["status"] = "error"
        rec["error_message"] = str(e)[:500]
        log.error("✗ %s: %s", key, e)
        return rec
    finally:
        rec["duration_s"] = round(time.time() - t0, 2)
        for p in (src, dst):
            if p and os.path.exists(p):
                os.unlink(p)


# Auditoria (parquet)
def _gravar_auditoria(s3, execution_id: str, registros: List[dict]) -> str:
    df = pd.DataFrame(registros)
    if "masked_columns" in df.columns:
        df["masked_columns"] = df["masked_columns"].apply(
            lambda v: json.dumps(v, ensure_ascii=False)
        )
    buf = io.BytesIO()
    df.to_parquet(buf, engine="pyarrow", index=False)
    buf.seek(0)
    key = f"{AUDIT_PREFIX}execution_id={execution_id}/part-0.parquet"
    s3.put_object(Bucket=MINIO_BUCKET, Key=key, Body=buf.getvalue())

    local = Path(__file__).parent / f"auditoria_mascaramento_{execution_id}.parquet"
    df.to_parquet(local, engine="pyarrow", index=False)
    log.info("Auditoria: s3://%s/%s (cópia local: %s)", MINIO_BUCKET, key, local)
    return key


# Main
def main() -> None:
    parser = argparse.ArgumentParser(description="Mascaramento de PII no raw/ do MinIO.")
    parser.add_argument("--apply", action="store_true",
                        help="Efetiva a sobrescrita em raw/. Sem esta flag roda em dry-run.")
    parser.add_argument("--force", action="store_true",
                        help="Reprocessa objetos já mascarados (tag masked=true).")
    parser.add_argument("--limit", type=int, default=0, help="Processa no máximo N objetos.")
    parser.add_argument("--pattern", default="", help="Filtra por substring na key.")
    parser.add_argument("--only-ext", default="",
                        help="Extensões a processar, ex.: csv,txt")
    parser.add_argument("--max-size-mb", type=int, default=0,
                        help="Pula objetos maiores que N MB (0 = sem limite). "
                             "Útil p/ fatiar dry-runs: pequenos primeiro, grandes depois.")
    parser.add_argument("--prefix", default=RAW_PREFIX, help="Prefixo a varrer (default raw/).")
    args = parser.parse_args()

    if not HMAC_SECRET:
        raise SystemExit(
            "MASKING_HMAC_SECRET não definida no .env — necessária para tokenizar CPF/NIS."
        )

    execution_id = uuid.uuid4().hex
    apply = args.apply
    only_ext = {("." + e.strip().lstrip(".")).lower() for e in args.only_ext.split(",") if e.strip()}

    log.info("=" * 70)
    log.info("Execução %s | modo=%s | prefixo=%s", execution_id,
             "APPLY (sobrescreve raw/)" if apply else "DRY-RUN (masked_dryrun/)", args.prefix)
    log.info("=" * 70)

    conn_str = _conn_str()
    _criar_control_table(conn_str)
    masked_hashes = set() if args.force else _carregar_masked_hashes(conn_str)

    s3 = _cliente_minio()

    registros: List[dict] = []
    contagem: Dict[str, int] = {}
    processados = 0

    for key, size in _listar_objetos(s3, args.prefix):
        if args.pattern and args.pattern not in key:
            continue
        ext = os.path.splitext(key)[1].lower()
        if only_ext and ext not in only_ext:
            continue
        # arquivos de lock/temporários do Excel (~$...) não são planilhas reais
        if os.path.basename(key).startswith("~$"):
            continue
        if args.max_size_mb and size > args.max_size_mb * 1024 * 1024:
            continue
        if args.limit and processados >= args.limit:
            break
        processados += 1

        if not args.force and _ja_mascarado(s3, key):
            rec = _novo_registro(execution_id, key)
            rec["status"] = "skipped_already"
            registros.append(rec)
            contagem["skipped_already"] = contagem.get("skipped_already", 0) + 1
            log.info("→ [%d] %s — já mascarado (tag), pulando", processados, key)
            continue

        rec = processar_objeto(s3, conn_str, key, execution_id, apply, masked_hashes)
        registros.append(rec)
        contagem[rec["status"]] = contagem.get(rec["status"], 0) + 1

        if rec["hash_before"] is not None:
            _registrar_control(conn_str, rec)

        icone = {"masked": "✓", "dry_run": "◐", "error": "✗"}.get(rec["status"], "·")
        log.info(
            "%s [%d] %s — %s | cols=%d | linhas=%d/%d",
            icone, processados, key, rec["status"],
            len(rec["masked_columns"]), rec["registros_alterados"], rec["registros_total"],
        )

    if registros:
        _gravar_auditoria(s3, execution_id, registros)

    log.info("=" * 70)
    log.info("Concluído. Objetos: %d", processados)
    for status, n in sorted(contagem.items()):
        log.info("  %-20s %d", status, n)
    if not apply:
        log.info("DRY-RUN — nada foi sobrescrito em raw/. Prévia em %s", DRYRUN_PREFIX)
    log.info("Log: %s", _LOG_FILE)


if __name__ == "__main__":
    main()
